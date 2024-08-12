#Reading Data from an Excel File:
from openpyxl import load_workbook

# Load the Excel file
workbook = load_workbook('Test.xlsx')

# Select the active worksheet
worksheet = workbook.active

# Access cell values
print(worksheet['A1'].value)  # Print value of cell A1
print(worksheet.cell(row=2, column=2).value)  # Print value of cell B2

# Iterate over rows and columns
for row in worksheet.iter_rows(min_row=1, max_row=4, min_col=1, max_col=2):
    for cell in row:
        print(cell.value, end='\t')
    print()  # Move to the next line
