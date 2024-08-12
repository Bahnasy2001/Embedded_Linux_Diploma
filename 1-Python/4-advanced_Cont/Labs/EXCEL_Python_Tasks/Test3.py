#Styling Cells:
from openpyxl import Workbook ,load_workbook
from openpyxl.styles import Font, Alignment

#Create a new Workbook
workbook = Workbook()

# Load the Excel file
workbook2 = load_workbook('Test.xlsx')

#Create a new WorkSheet
worksheet = workbook.active
worksheet2 = workbook2.active

#write data to cells
worksheet['A1'] = 'Bold and Centered'
worksheet['A1'].font = Font(bold=True)
worksheet['A1'].alignment = Alignment(horizontal='center')

worksheet2['A1'].font = Font(bold=True)
worksheet2['A1'].alignment = Alignment(horizontal='center')

# Save the workbook
workbook.save('styled_output.xlsx')
# Save the workbook
workbook2.save('Test.xlsx')